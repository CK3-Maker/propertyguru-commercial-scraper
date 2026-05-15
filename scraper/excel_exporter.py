"""Export scrape results to formatted Excel workbook."""

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from scraper.config import REPORTS_DIR

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _auto_width(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_excel(
    new_listings: list[dict],
    summary: dict,
    now: datetime,
) -> Path:
    """Create the Excel report and return its path."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"propertyguru_new_listings_{now.strftime('%Y-%m-%d_%H%M')}_MY.xlsx"
    filepath = REPORTS_DIR / filename

    # --- Sheet 1: Summary ---
    summary_rows = [
        {"Field": k, "Value": v}
        for k, v in summary.items()
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Sheet 2: New Listings ---
    new_cols = [
        "No.", "Source", "Listing ID", "Title", "URL", "Price", "Price Note", "Price Numeric",
        "Location", "State", "Property Type", "Tenure", "Furnishing",
        "Land Area", "Price Per Sqft",
        "Agent Name", "Agency Name", "Posted Date",
        "Description", "Commercial Keywords",
        "Main Image URL", "Photos", "First Seen At",
    ]
    new_rows = []
    for i, lst in enumerate(new_listings, 1):
        kw = lst.get("commercial_keywords", [])
        # Use detail-page land_area, fall back to card land_size
        land = lst.get("land_area") or lst.get("land_size", "")
        # Use detail-page psf if available
        psf = lst.get("psf_detail") or lst.get("psf", "")
        new_rows.append({
            "No.": i,
            "Source": lst.get("source", ""),
            "Listing ID": lst.get("listing_id_page") or lst.get("listing_id", ""),
            "Title": lst.get("detail_title") or lst.get("title", ""),
            "URL": lst.get("url", ""),
            "Price": lst.get("price", ""),
            "Price Note": lst.get("price_note", ""),
            "Price Numeric": lst.get("price_numeric"),
            "Location": lst.get("address") or lst.get("location", ""),
            "State": lst.get("state", ""),
            "Property Type": lst.get("property_type", ""),
            "Tenure": lst.get("tenure", ""),
            "Furnishing": lst.get("furnishing", ""),
            "Land Area": land,
            "Price Per Sqft": psf,
            "Agent Name": lst.get("agent_name_detail") or lst.get("agent_name", ""),
            "Agency Name": lst.get("agency_name_detail") or lst.get("agency_name", ""),
            "Posted Date": lst.get("posted_date") or lst.get("posted_date_text", ""),
            "Description": lst.get("description", ""),
            "Commercial Keywords": ", ".join(kw) if kw else "",
            "Main Image URL": lst.get("main_image_url", ""),
            "Photos": lst.get("photo_count", 0),
            "First Seen At": lst.get("first_seen_at", ""),
        })
    df_new = pd.DataFrame(new_rows, columns=new_cols) if new_rows else pd.DataFrame(columns=new_cols)

    # --- Sheet 3: Raw Data ---
    if new_listings:
        # Remove internal fields for cleaner raw output
        clean_listings = [
            {k: v for k, v in lst.items() if not k.startswith("_")}
            for lst in new_listings
        ]
        df_raw = pd.json_normalize(clean_listings)
    else:
        df_raw = pd.DataFrame()

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_new.to_excel(writer, sheet_name="New Listings", index=False)
        df_raw.to_excel(writer, sheet_name="Raw Data", index=False)

    wb = load_workbook(filepath)

    # Style Summary
    ws_sum = wb["Summary"]
    _style_header(ws_sum)
    _auto_width(ws_sum)

    # Style New Listings
    ws_new = wb["New Listings"]
    _style_header(ws_new)

    # Wrap description columns
    desc_cols = {"Description", "Commercial Keywords"}
    header_map = {cell.value: cell.column for cell in ws_new[1] if cell.value}
    for col_name in desc_cols:
        if col_name in header_map:
            col_idx = header_map[col_name]
            for row in ws_new.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Hyperlinks for URL column
    url_col = header_map.get("URL")
    if url_col:
        for row in ws_new.iter_rows(min_row=2, min_col=url_col, max_col=url_col):
            for cell in row:
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = str(cell.value)
                    cell.font = Font(color="0563C1", underline="single")

    # Highlight rows with missing price
    price_col = header_map.get("Price")
    if price_col:
        for row in ws_new.iter_rows(min_row=2, min_col=1, max_col=ws_new.max_column):
            price_cell = row[price_col - 1]
            if not price_cell.value or str(price_cell.value).strip() == "":
                for cell in row:
                    cell.fill = WARN_FILL

    _auto_width(ws_new)

    # Style Raw Data
    ws_raw = wb["Raw Data"]
    _style_header(ws_raw)
    _auto_width(ws_raw, max_width=40)

    wb.save(filepath)
    logger.info("Excel report saved: %s", filepath)
    return filepath
